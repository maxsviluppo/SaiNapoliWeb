import sys

filepath = r'C:\Users\Max\Downloads\DENTISTI. SI EXEL.xls'

# Try reading as xlsx zip first (in case it is xlsx format with .xls extension)
import zipfile
is_zip = zipfile.is_zipfile(filepath)
print("Is zip/xlsx:", is_zip)

if is_zip:
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    print("Sheets in workbook:", wb.sheetnames)
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"\n--- Sheet: {name} ({ws.max_row} rows, {ws.max_column} cols) ---")
        for r in range(1, min(15, ws.max_row + 1)):
            row_vals = [ws.cell(r, c).value for c in range(1, min(25, ws.max_column + 1))]
            if any(v is not None for v in row_vals):
                print(f"Row {r}: {row_vals}")
else:
    # Try xlrd
    try:
        import xlrd
        wb = xlrd.open_workbook(filepath)
        print("Sheets in xlrd workbook:", wb.sheet_names())
        for name in wb.sheet_names():
            ws = wb.sheet_by_name(name)
            print(f"\n--- Sheet: {name} ({ws.nrows} rows, {ws.ncols} cols) ---")
            for r in range(min(15, ws.nrows)):
                print(f"Row {r+1}: {ws.row_values(r)[:25]}")
    except Exception as e:
        print("xlrd error:", e)
